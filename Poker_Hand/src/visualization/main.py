import matplotlib.pyplot as plt
import matplotlib.patches as patches
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image

def create_range_chart(range_hands, title, output_path):
    """
    Creates a poker hand range chart that visually highlights specific hands and saves it.
    Args:
        range_hands (set): A set of strings for the hands to highlight (e.g., {'AA', 'KK', 'AKs'}).
        title (str): The title to display above the chart.
        output_path (str): Path to save the PNG file.
    """
    # --- 1. Setup the Chart and Ranking Order ---
    ranks = ['A', 'K', 'Q', 'J', 'T', '9', '8', '7', '6', '5', '4', '3', '2']
    
    # Use a dark background for the figure for better contrast
    fig, ax = plt.subplots(figsize=(10, 10))
    fig.patch.set_facecolor('#2d2d2d')  # Dark grey background
    ax.set_facecolor('#2d2d2d')
    
    # --- 2. Iterate Through Grid and Draw Colored Cells ---
    for i, rank1 in enumerate(ranks):
        for j, rank2 in enumerate(ranks):
            # Determine the hand name based on its position in the grid
            if i == j:  # Pocket Pairs (diagonal)
                hand = rank1 + rank2
            elif j > i:  # Suited Hands (upper right)
                hand = rank1 + rank2 + 's'
            else:       # Offsuit Hands (lower left)
                hand = rank2 + rank1 + 'o'
            
            # Determine the color based on whether the hand is in the specified range
            is_in_range = hand in range_hands
            
            # Use specific hex colors to match the example image
            cell_color = '#d62f2f' if is_in_range else '#696969'  # Red or Grey
            text_color = 'white'
            
            # Draw the cell rectangle
            ax.add_patch(
                patches.Rectangle(
                    (j, 12 - i),  # (x, y)
                    1, 1,         # width, height
                    facecolor=cell_color,
                    edgecolor='#404040',  # Border color for cells
                    linewidth=2
                )
            )
            
            # Add the hand text in the center of the cell
            ax.text(j + 0.5, 12 - i + 0.5, hand,
                    ha='center', va='center',
                    color=text_color,
                    fontsize=14,
                    fontweight='bold')
    
    # --- 3. Finalize Chart Appearance ---
    ax.set_xlim(0, 13)
    ax.set_ylim(0, 13)
    ax.set_aspect('equal', adjustable='box')
    
    # Remove axis ticks and labels
    ax.set_xticks([])
    ax.set_yticks([])
    
    # Set the main title for the chart
    ax.set_title(title, color='white', fontsize=28, fontweight='bold', pad=20)
    
    plt.tight_layout()
    
    # Save the chart instead of showing it
    plt.savefig(output_path, dpi=150, bbox_inches='tight', facecolor='#2d2d2d')
    plt.close()  # Close to free memory

def load_range_from_csv(csv_path):
    """
    Load poker hand range from CSV file.
    Args:
        csv_path (str): Path to the CSV file
    Returns:
        set: Set of hands where in_range=True
    """
    try:
        df = pd.read_csv(csv_path)
        # Filter hands where in_range is True
        in_range_hands = df[df['in_range'] == True]['hand'].tolist()
        return set(in_range_hands)
    except Exception as e:
        print(f"Error loading CSV {csv_path}: {e}")
        return set()

def create_excel_workbook(table_size, positions, actions, base_path, output_path, charts_path):
    """
    Create Excel workbook with sheets for each position containing range data and embedded PNG charts.
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    for position in positions:
        ws = wb.create_sheet(title=position)
        
        # Create headers
        headers = ['Hand'] + [action.replace('_', '-').title() for action in actions]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Generate all possible hands
        ranks = ['A', 'K', 'Q', 'J', 'T', '9', '8', '7', '6', '5', '4', '3', '2']
        all_hands = []
        
        # Pairs
        for rank in ranks:
            all_hands.append(rank + rank)
        
        # Suited and offsuit hands
        for i in range(len(ranks)):
            for j in range(i + 1, len(ranks)):
                all_hands.append(ranks[i] + ranks[j] + 's')
                all_hands.append(ranks[i] + ranks[j] + 'o')
        
        all_hands.sort()
        
        # Fill in hand data
        for row_num, hand in enumerate(all_hands, 2):
            ws.cell(row=row_num, column=1, value=hand)
            
            for col_num, action in enumerate(actions, 2):
                csv_file = f"{base_path}/{table_size}_player/{position}/{action}/low_winrate_hands.csv"
                
                if os.path.exists(csv_file):
                    range_hands = load_range_from_csv(csv_file)
                    in_range = "YES" if hand in range_hands else "NO"
                    
                    cell = ws.cell(row=row_num, column=col_num, value=in_range)
                    if in_range == "YES":
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                else:
                    ws.cell(row=row_num, column=col_num, value="N/A")
        
        # Add PNG charts to the right of the data table
        chart_start_col = len(headers) + 2  # Start 2 columns after the data
        chart_row = 2  # Start from row 2
        
        for i, action in enumerate(actions):
            # Construct the PNG file path
            table_suffix = "6max" if table_size == "6" else "9max"
            png_file = f"{charts_path}/{position}_{action}_{table_suffix}.png"
            
            if os.path.exists(png_file):
                # Add a header for the chart
                header_cell = ws.cell(row=chart_row - 1, column=chart_start_col + i * 15, 
                                    value=f"{action.replace('_', '-').title()} Chart")
                header_cell.font = Font(bold=True, size=14)
                
                # Load and embed the image
                img = Image(png_file)
                
                # Resize image to fit nicely in Excel (adjust size as needed)
                img.width = 400  # pixels
                img.height = 400  # pixels
                
                # Position the image
                img.anchor = f"{chr(ord('A') + chart_start_col + i * 15 - 1)}{chart_row}"
                ws.add_image(img)
        
        # Adjust column widths for better visibility
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(ord('A') + col - 1)].width = 12
    
    wb.save(output_path)

if __name__ == '__main__':
    print("Loading poker hand ranges from CSV data...")
    
    # Base paths
    base_path = "Poker_Hand/output"
    output_base = r"C:\Users\cinco\Desktop\Poker_Analysis\Poker_Hand\src\visualization\output"
    
    # Create output directories
    os.makedirs(f"{output_base}/6_player/charts", exist_ok=True)
    os.makedirs(f"{output_base}/9_player/charts", exist_ok=True)
    
    # Define positions for each table size
    positions_6max = ["UTG", "HJ", "CO", "BTN", "SB", "BB"]
    positions_9max = ["UTG", "UTG1", "MP1", "MP2", "HJ", "CO", "BTN", "SB", "BB"]
    
    # Define actions to visualize
    actions = ["opening_raise", "3_bet", "4_bet"]
    
    try:
        print("=" * 60)
        print("GENERATING COMPREHENSIVE POKER RANGE VISUALIZATIONS")
        print("=" * 60)
        
        # === 6-MAX CHARTS ===
        print("\n" + "="*30)
        print("6-MAX RANGES")
        print("="*30)
        
        for action in actions:
            print(f"\n--- 6-MAX {action.upper().replace('_', '-')} RANGES ---")
            for position in positions_6max:
                csv_file = f"{base_path}/6_player/{position}/{action}/low_winrate_hands.csv"
                
                if os.path.exists(csv_file):
                    print(f"Creating {position} {action.replace('_', '-')} range (6-max)...")
                    range_hands = load_range_from_csv(csv_file)
                    title = f"{position} {action.replace('_', '-').title()} Range (6-max)"
                    output_path = f"{output_base}/6_player/charts/{position}_{action}_6max.png"
                    create_range_chart(range_hands, title, output_path)
                else:
                    print(f"Skipping {position} {action} (6-max) - file not found")
        
        # === 9-MAX CHARTS ===
        print("\n" + "="*30)
        print("9-MAX RANGES") 
        print("="*30)
        
        for action in actions:
            print(f"\n--- 9-MAX {action.upper().replace('_', '-')} RANGES ---")
            for position in positions_9max:
                csv_file = f"{base_path}/9_player/{position}/{action}/low_winrate_hands.csv"
                
                if os.path.exists(csv_file):
                    print(f"Creating {position} {action.replace('_', '-')} range (9-max)...")
                    range_hands = load_range_from_csv(csv_file)
                    title = f"{position} {action.replace('_', '-').title()} Range (9-max)"
                    output_path = f"{output_base}/9_player/charts/{position}_{action}_9max.png"
                    create_range_chart(range_hands, title, output_path)
                else:
                    print(f"Skipping {position} {action} (9-max) - file not found")
        
        # === CREATE EXCEL FILES ===
        print("\n" + "="*30)
        print("CREATING EXCEL FILES")
        print("="*30)
        
        print("Creating 6-max Excel workbook...")
        excel_6max = f"{output_base}/6_player/6max_ranges.xlsx"
        charts_6max = f"{output_base}/6_player/charts"
        create_excel_workbook("6", positions_6max, actions, base_path, excel_6max, charts_6max)
        
        print("Creating 9-max Excel workbook...")
        excel_9max = f"{output_base}/9_player/9max_ranges.xlsx"
        charts_9max = f"{output_base}/9_player/charts"
        create_excel_workbook("9", positions_9max, actions, base_path, excel_9max, charts_9max)
        
        print("\n" + "="*60)
        print("ALL FILES GENERATED SUCCESSFULLY!")
        print("="*60)
        print(f"Output saved to: {output_base}")
        print("Generated files:")
        print("• PNG charts for all position/action combinations")
        print("• 6max_ranges.xlsx - Excel file with all 6-max ranges")
        print("• 9max_ranges.xlsx - Excel file with all 9-max ranges")
        print("• Each Excel sheet represents a position with YES/NO for each hand in each action")
        
    except Exception as e:
        print(f"Error occurred: {e}")
        print("Make sure pandas, matplotlib, and openpyxl are installed:")
        print("pip install pandas matplotlib openpyxl")
